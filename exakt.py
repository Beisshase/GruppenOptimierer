import time
import openpyxl
import pulp

EXCEL_OUT = "Abstandsmatrix.xlsx"
N_GRUPPEN = 4
ZEITLIMIT_SEKUNDEN = 120

# ----------------------------------------------------------------------
# Matrix aus vorhandener Abstandsmatrix.xlsx laden
# ----------------------------------------------------------------------
wb = openpyxl.load_workbook(EXCEL_OUT, data_only=True)
sh = wb["Abstandsmatrix_km"]
N = sh.max_column - 1
labels = [sh.cell(row=1, column=2 + j).value for j in range(N)]
matrix = [[sh.cell(row=2 + i, column=2 + j).value for j in range(N)] for i in range(N)]
valid = [any(v is not None for v in matrix[i]) for i in range(N)]
gueltig = [i for i in range(N) if valid[i]]
M = len(gueltig)

def dist(i, j):
    a, b = matrix[i][j], matrix[j][i]
    if a is None and b is None: return 0.0
    if a is None: return b
    if b is None: return a
    return (a + b) / 2

basis, rest = M // N_GRUPPEN, M % N_GRUPPEN
groessen = [basis + (1 if g < rest else 0) for g in range(N_GRUPPEN)]
print(f"{M} Vereine, Zielgroessen je Gruppe: {groessen}")

paare = [(a, b) for ai, a in enumerate(gueltig) for b in gueltig[ai + 1:]]
print(f"{len(paare)} Vereinspaare x {N_GRUPPEN} Gruppen = {len(paare) * N_GRUPPEN} Kopplungsvariablen")

# ----------------------------------------------------------------------
# MILP: x[i,g] = 1 wenn Verein i in Gruppe g; z[a,b,g] = 1 wenn a UND b in Gruppe g
# ----------------------------------------------------------------------
prob = pulp.LpProblem("Gruppenoptimierung", pulp.LpMinimize)

x = {(i, g): pulp.LpVariable(f"x_{i}_{g}", cat="Binary")
     for i in gueltig for g in range(N_GRUPPEN)}
z = {(a, b, g): pulp.LpVariable(f"z_{a}_{b}_{g}", lowBound=0, upBound=1)
     for (a, b) in paare for g in range(N_GRUPPEN)}

for i in gueltig:
    prob += pulp.lpSum(x[i, g] for g in range(N_GRUPPEN)) == 1

for g in range(N_GRUPPEN):
    prob += pulp.lpSum(x[i, g] for i in gueltig) == groessen[g]

for (a, b) in paare:
    for g in range(N_GRUPPEN):
        prob += z[a, b, g] <= x[a, g]
        prob += z[a, b, g] <= x[b, g]
        prob += z[a, b, g] >= x[a, g] + x[b, g] - 1

prob += pulp.lpSum(dist(a, b) * z[a, b, g] for (a, b) in paare for g in range(N_GRUPPEN))

solver = pulp.PULP_CBC_CMD(timeLimit=ZEITLIMIT_SEKUNDEN, msg=True)
start = time.time()
status = prob.solve(solver)
dauer = time.time() - start

print(f"\nStatus: {pulp.LpStatus[status]}, Dauer: {dauer:.1f}s")
print(f"Zielwert (Distanzsumme): {pulp.value(prob.objective):.1f} km")

gruppen = [[] for _ in range(N_GRUPPEN)]
for i in gueltig:
    for g in range(N_GRUPPEN):
        if pulp.value(x[i, g]) > 0.5:
            gruppen[g].append(i)

for gi, grp in enumerate(gruppen, 1):
    summe = sum(dist(a, b) for ai, a in enumerate(grp) for b in grp[ai + 1:])
    print(f"Gruppe {gi} ({len(grp)} Vereine, {summe:.1f} km intern): " + ", ".join(labels[i] for i in grp))
