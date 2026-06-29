import time
import openpyxl
from ortools.sat.python import cp_model

EXCEL_OUT = "Abstandsmatrix.xlsx"
N_GRUPPEN = 4
ZEITLIMIT_SEKUNDEN = 1800

# ----------------------------------------------------------------------
# Matrix aus vorhandener Abstandsmatrix.xlsx laden
# ----------------------------------------------------------------------
wb = openpyxl.load_workbook(EXCEL_OUT, data_only=True)
sh = wb["Abstandsmatrix_km"]
N = sh.max_column - 1
labels = [sh.cell(row=1, column=2 + j).value for j in range(N)]
matrix = [[sh.cell(row=2 + i, column=2 + j).value for j in range(N)] for i in range(N)]
valid = [any(v is not None for v in matrix[i]) for i in range(N)]
gueltig = [i for i in range(N) if valid[i]]
M = len(gueltig)

def dist(i, j):
    a, b = matrix[i][j], matrix[j][i]
    if a is None and b is None: return 0.0
    if a is None: return b
    if b is None: return a
    return (a + b) / 2

basis, rest = M // N_GRUPPEN, M % N_GRUPPEN
groessen = [basis + (1 if g < rest else 0) for g in range(N_GRUPPEN)]
print(f"{M} Vereine, Zielgroessen je Gruppe: {groessen}")

paare = [(a, b) for ai, a in enumerate(gueltig) for b in gueltig[ai + 1:]]
print(f"{len(paare)} Vereinspaare x {N_GRUPPEN} Gruppen = {len(paare) * N_GRUPPEN} Kopplungsvariablen")

# Distanzen in Zehntel-km als Integer (CP-SAT braucht Integer-Koeffizienten)
def dist_int(i, j):
    return round(dist(i, j) * 10)

# ----------------------------------------------------------------------
# CP-SAT Modell: x[i,g] = 1 wenn Verein i in Gruppe g; z[a,b,g] = 1 wenn a UND b in Gruppe g
# ----------------------------------------------------------------------
model = cp_model.CpModel()

x = {(i, g): model.NewBoolVar(f"x_{i}_{g}") for i in gueltig for g in range(N_GRUPPEN)}

for i in gueltig:
    model.AddExactlyOne(x[i, g] for g in range(N_GRUPPEN))

for g in range(N_GRUPPEN):
    model.Add(sum(x[i, g] for i in gueltig) == groessen[g])

z = {}
for (a, b) in paare:
    for g in range(N_GRUPPEN):
        zv = model.NewBoolVar(f"z_{a}_{b}_{g}")
        model.AddBoolAnd([x[a, g], x[b, g]]).OnlyEnforceIf(zv)
        model.AddBoolOr([x[a, g].Not(), x[b, g].Not()]).OnlyEnforceIf(zv.Not())
        z[a, b, g] = zv

model.Minimize(sum(dist_int(a, b) * z[a, b, g] for (a, b) in paare for g in range(N_GRUPPEN)))

# Heuristik-Loesung (Blatt "Gruppen") als Hint fuer den Solver, falls vorhanden
if "Gruppen" in wb.sheetnames:
    sh_gr = wb["Gruppen"]
    vnr_zu_gruppe = {}
    for row in sh_gr.iter_rows(min_row=2, max_col=2, values_only=True):
        gi, vnr = row
        if gi is None or vnr is None:
            continue
        vnr_zu_gruppe[str(vnr)] = int(gi) - 1
    anzahl_hints = 0
    for i in gueltig:
        zugew = vnr_zu_gruppe.get(str(labels[i]))
        if zugew is None:
            continue
        for g in range(N_GRUPPEN):
            model.add_hint(x[i, g], 1 if g == zugew else 0)
        anzahl_hints += 1
    if anzahl_hints:
        print(f"Hint aus Blatt 'Gruppen' fuer {anzahl_hints} Vereine gesetzt.")

solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = ZEITLIMIT_SEKUNDEN
solver.parameters.num_search_workers = 16
solver.parameters.log_search_progress = True

start = time.time()
status = solver.Solve(model)
dauer = time.time() - start

print(f"\nStatus: {solver.StatusName(status)}, Dauer: {dauer:.1f}s")
print(f"Zielwert (Distanzsumme): {solver.ObjectiveValue() / 10:.1f} km")
print(f"Beste bekannte untere Schranke: {solver.BestObjectiveBound() / 10:.1f} km")

gruppen = [[] for _ in range(N_GRUPPEN)]
for i in gueltig:
    for g in range(N_GRUPPEN):
        if solver.Value(x[i, g]):
            gruppen[g].append(i)

for gi, grp in enumerate(gruppen, 1):
    summe = sum(dist(a, b) for ai, a in enumerate(grp) for b in grp[ai + 1:])
    print(f"Gruppe {gi} ({len(grp)} Vereine, {summe:.1f} km intern): " + ", ".join(labels[i] for i in grp))
