import os
import re
import time
import random
import numpy as np
import openpyxl
import requests
from openlocationcode import openlocationcode as olc

EXCEL_IN   = "Meldelisten-mit-Sportstätten_A1.xlsx"
_basis, _ext = os.path.splitext(EXCEL_IN)
EXCEL_OUT  = f"{_basis}-Gruppeneinteilung{_ext}"
SHEET      = "AZ"
N_GRUPPEN  = 4          # <- hier später variieren
SEED       = 42
VERSUCHE   = 40         # Zufallsneustarts fuer die Heuristik
OSRM_BATCH = 50         # max Koordinaten pro OSRM-Anfrage; Inter-Chunk-Requests haben 2x davon

# ----------------------------------------------------------------------
# 1) Excel einlesen: Spalte A = V.Nr., Spalte B = Vereinsname, Spalte L = Adresse
# ----------------------------------------------------------------------
wb = openpyxl.load_workbook(EXCEL_IN, data_only=True)
ws = wb[SHEET]

vereine = []
for row in ws.iter_rows(min_row=2, values_only=True):
    vnr = row[0]
    name = row[1]
    adr = row[11]
    if vnr is None:
        continue
    name = str(name).strip() if name else ""
    adr = str(adr).strip() if adr else ""
    vereine.append((str(vnr), name, adr))

print(f"{len(vereine)} Vereine eingelesen.")

labels = [vnr for vnr, _, _ in vereine]
namen = [name for _, name, _ in vereine]
N = len(vereine)

# ----------------------------------------------------------------------
# 2) Adresse bereinigen: Adresse besteht aus "Straße, PLZ Ort" oder
#    "Pluscode, PLZ Ort"
# ----------------------------------------------------------------------
PLUSCODE_REGEX = re.compile(r"\b[23456789CFGHJMPQRVWXcfghjmpqrvwx]{2,8}\+[23456789CFGHJMPQRVWXcfghjmpqrvwx]{0,3}\b")

def parse_adresse(adresse):
    # Adresse ist bereinigt auf "Strasse/Platzname, PLZ Ort" bzw. "Pluscode,
    # PLZ Ort" - die Strasse muss daher kein bestimmtes Wort (Weg/Strasse/...)
    # enthalten, sie ist einfach das Segment, das weder PLZ/Ort noch Pluscode ist.
    s = adresse.replace("  ", " ").strip()
    teile = [t.strip() for t in s.split(",") if t.strip()]

    plz_ort = None
    plz_index = None
    for idx, t in enumerate(teile):
        m = re.search(r"\b(\d{5})\s+(.+)", t)
        if m:
            ort = re.split(r"[-]| Zentrum| Innenstadt| Stadtgebiet| Süd| Nord| Ost| West",
                           m.group(2))[0].strip()
            plz_ort = f"{m.group(1)} {ort}"
            plz_index = idx
            break

    pluscode = None
    pluscode_index = None
    for idx, t in enumerate(teile):
        for token in PLUSCODE_REGEX.findall(t):
            if olc.isValid(token):
                pluscode = token.upper()
                pluscode_index = idx
                break
        if pluscode:
            break

    strasse = None
    for idx, t in enumerate(teile):
        if idx == plz_index or idx == pluscode_index:
            continue
        if re.search(r"\boder\b", t, re.I):
            # eher eine unsichere Ortsbeschreibung ("Kemter Wiesen oder Mainzer
            # Str. 199") als eine echte Adresse - nicht als Strasse verwenden.
            continue
        strasse = t
        break

    return strasse, pluscode, plz_ort

# ----------------------------------------------------------------------
# 3) Geocoding (Photon, mit Pluscode-Unterstuetzung)
# ----------------------------------------------------------------------
def photon_suche(query):
    headers = {"User-Agent": "spielfeld-abstandsmatrix/1.0"}
    r = requests.get("https://photon.komoot.io/api/",
                     params={"q": query, "limit": 1, "lang": "de"},
                     headers=headers, timeout=30)
    r.raise_for_status()
    feats = r.json().get("features", [])
    time.sleep(1.0)
    if feats:
        lon, lat = feats[0]["geometry"]["coordinates"]
        return (float(lat), float(lon))
    return None

def geocode(adresse, cache={}):
    # quelle gibt an, wie praezise die Koordinate ist: "pluscode"/"strasse"
    # sind genau, "ortsmitte" (nur PLZ/Ort) und "rohtext" sind Fallbacks.
    if adresse in cache:
        return cache[adresse]
    strasse, pluscode, plz_ort = parse_adresse(adresse)
    result = None
    quelle = None

    if pluscode and plz_ort:
        referenz = photon_suche(plz_ort)
        if referenz:
            try:
                voller_code = olc.recoverNearest(pluscode, referenz[0], referenz[1])
                bereich = olc.decode(voller_code)
                result = (bereich.latitudeCenter, bereich.longitudeCenter)
                quelle = "pluscode"
            except Exception:
                result = None

    if result is None:
        varianten = []
        if strasse and plz_ort:
            varianten.append(("strasse", f"{strasse}, {plz_ort}"))
        if plz_ort:
            varianten.append(("ortsmitte", plz_ort))
        if not varianten:
            varianten.append(("rohtext", adresse))
        for label, variante in varianten:
            result = photon_suche(variante)
            if result:
                quelle = label
                break
    cache[adresse] = (result, quelle)
    return cache[adresse]


def osrm_matrix_chunked(gueltige_coords, on_block=None):
    """Berechnet MxM-Distanzmatrix in Batches (umgeht URL-Laengenlimit bei >OSRM_BATCH Coords)."""
    M = len(gueltige_coords)
    if M <= OSRM_BATCH:
        coord_str = ";".join(f"{lon:.6f},{lat:.6f}" for lat, lon in gueltige_coords)
        r = requests.get(f"https://router.project-osrm.org/table/v1/driving/{coord_str}",
                         params={"annotations": "distance"}, timeout=120)
        r.raise_for_status()
        osrm = r.json()["distances"]
        if on_block:
            on_block(1, 1)
        else:
            print()
        return [[round(osrm[a][b] / 1000, 1) if osrm[a][b] is not None else None
                 for b in range(M)] for a in range(M)]

    raw = [[None] * M for _ in range(M)]
    chunks = [list(range(s, min(s + OSRM_BATCH, M))) for s in range(0, M, OSRM_BATCH)]
    n_chunks = len(chunks)
    n_tasks = n_chunks * (n_chunks + 1) // 2
    fertig = 0
    for ci, ca in enumerate(chunks):
        for cj in range(ci, n_chunks):
            cb = chunks[cj]
            combined = ca if ci == cj else ca + cb
            n_a = len(ca)
            coord_str = ";".join(f"{gueltige_coords[k][1]:.6f},{gueltige_coords[k][0]:.6f}"
                                 for k in combined)
            r = requests.get(f"https://router.project-osrm.org/table/v1/driving/{coord_str}",
                             params={"annotations": "distance"}, timeout=120)
            r.raise_for_status()
            osrm = r.json()["distances"]
            if ci == cj:
                for a, i in enumerate(ca):
                    for b, j in enumerate(ca):
                        d = osrm[a][b]
                        raw[i][j] = round(d / 1000, 1) if d is not None else None
            else:
                for a, i in enumerate(ca):
                    for b, j in enumerate(cb):
                        d_ab = osrm[a][n_a + b]
                        d_ba = osrm[n_a + b][a]
                        raw[i][j] = round(d_ab / 1000, 1) if d_ab is not None else None
                        raw[j][i] = round(d_ba / 1000, 1) if d_ba is not None else None
            fertig += 1
            if on_block:
                on_block(fertig, n_tasks)
            else:
                print(f"  OSRM-Matrix: Block {fertig}/{n_tasks} ", end="\r", flush=True)
    if not on_block:
        print()
    return raw


nicht_gefunden = []
ungenau = []

if os.path.exists(EXCEL_OUT):
    print(f"\n{EXCEL_OUT} existiert bereits -> Matrix wird daraus geladen (kein erneutes Geocoding/OSRM).")
    cache_wb = openpyxl.load_workbook(EXCEL_OUT, data_only=True)
    cache_sh = cache_wb["Abstandsmatrix_km"]
    cached_labels = [cache_sh.cell(row=1, column=2 + j).value for j in range(N)]
    if cached_labels != labels:
        raise SystemExit(
            f"{EXCEL_OUT} passt nicht zur aktuellen Vereinsliste in {EXCEL_IN}. "
            f"Bitte {EXCEL_OUT} löschen, um die Matrix neu zu berechnen."
        )
    matrix = [[cache_sh.cell(row=2 + i, column=2 + j).value for j in range(N)] for i in range(N)]
    valid = [any(v is not None for v in matrix[i]) for i in range(N)]
    for (vnr, _, adr), ok in zip(vereine, valid):
        if not ok:
            nicht_gefunden.append((vnr, adr if adr else "(keine Adresse)"))
else:
    print(f"\n{EXCEL_OUT} noch nicht vorhanden -> Matrix wird neu berechnet (Geocoding + OSRM).")
    coords = []
    for vnr, _, adr in vereine:
        if not adr:
            coords.append(None)
            nicht_gefunden.append((vnr, "(keine Adresse)"))
            print(f"{vnr}: KEINE Adresse")
            continue
        c, q = geocode(adr)
        coords.append(c)
        if c is None:
            nicht_gefunden.append((vnr, adr))
            print(f"{vnr}: NICHT gefunden -> {adr}")
        else:
            if q in ("ortsmitte", "rohtext"):
                ungenau.append((vnr, adr))
            print(f"{vnr}: {c[0]:.5f}, {c[1]:.5f}" + ("  [nur Ortsmitte/PLZ]" if q in ("ortsmitte", "rohtext") else ""))

    # ------------------------------------------------------------------
    # 4) Fahrtkilometer-Matrix (OSRM)
    # ------------------------------------------------------------------
    gueltig_idx = [i for i, c in enumerate(coords) if c is not None]
    gueltige_coords = [coords[i] for i in gueltig_idx]

    print(f"Berechne Fahrtkilometer-Matrix (OSRM, {len(gueltige_coords)} Vereine)...")
    raw = osrm_matrix_chunked(gueltige_coords)
    matrix = [[None] * N for _ in range(N)]
    for a, i in enumerate(gueltig_idx):
        for b, j in enumerate(gueltig_idx):
            matrix[i][j] = raw[a][b]

    valid = [c is not None for c in coords]

# Symmetrische Distanzmatrix fuer vektorisierte Optimierung (Ø beider Richtungen)
sym = np.zeros((N, N))
for _i in range(N):
    for _j in range(N):
        _a, _b = matrix[_i][_j], matrix[_j][_i]
        if _a is not None and _b is not None:
            sym[_i, _j] = (_a + _b) / 2.0
        elif _a is not None:
            sym[_i, _j] = float(_a)
        elif _b is not None:
            sym[_i, _j] = float(_b)

# ----------------------------------------------------------------------
# 5) Matrix in Excel (Blatt 1)
# ----------------------------------------------------------------------
out = openpyxl.Workbook()
sh = out.active
sh.title = "Abstandsmatrix_km"
sh.cell(row=1, column=1, value="V.Nr.")
for j, lab in enumerate(labels):
    sh.cell(row=1, column=2 + j, value=lab)
for i, lab in enumerate(labels):
    sh.cell(row=2 + i, column=1, value=lab)
    for j in range(N):
        sh.cell(row=2 + i, column=2 + j, value=matrix[i][j])

# ----------------------------------------------------------------------
# 6) In N_GRUPPEN ausgewogene Gruppen aufteilen (minimale interne Distanz)
# ----------------------------------------------------------------------
random.seed(SEED)
gueltig = [i for i in range(N) if valid[i]]
M = len(gueltig)

def dist(i, j):
    a, b = matrix[i][j], matrix[j][i]
    if a is None and b is None: return 0.0
    if a is None: return b
    if b is None: return a
    return (a + b) / 2

basis, rest = M // N_GRUPPEN, M % N_GRUPPEN
groessen = [basis + (1 if g < rest else 0) for g in range(N_GRUPPEN)]

def gruppen_kosten(m):
    s = 0.0
    for x in range(len(m)):
        for y in range(x + 1, len(m)):
            s += dist(m[x], m[y])
    return s

def gesamt_kosten(gr): return sum(gruppen_kosten(g) for g in gr)

def startloesung():
    pool = gueltig[:]
    random.shuffle(pool)
    gr, pos = [], 0
    for g in range(N_GRUPPEN):
        gr.append(pool[pos:pos + groessen[g]])
        pos += groessen[g]
    return gr

def optimiere(gr):
    # Pro Gruppenpaar wird der beste Tausch vektorisiert berechnet (numpy-Delta-Formel).
    # Kostendelta fuer Tausch a<->b: sA[b]-sA[a] + sB[a]-sB[b] - 2*sym[a,b]
    # wobei sA[x] = sum(sym[x, m] for m in Gruppe_A), analog sB.
    verbessert = True
    while verbessert:
        verbessert = False
        for ga in range(N_GRUPPEN):
            for gb in range(ga + 1, N_GRUPPEN):
                A = np.array(gr[ga])
                B = np.array(gr[gb])
                sA = sym[:, A].sum(axis=1)   # sA[x] = Σ sym[x, a] fuer a in A
                sB = sym[:, B].sum(axis=1)   # sB[x] = Σ sym[x, b] fuer b in B
                delta = (sA[B][np.newaxis, :] - sA[A][:, np.newaxis]
                       + sB[A][:, np.newaxis] - sB[B][np.newaxis, :]
                       - 2.0 * sym[np.ix_(A, B)])
                best = int(delta.argmin())
                best_ia, best_ib = divmod(best, len(B))
                if delta.flat[best] < -1e-9:
                    gr[ga][best_ia], gr[gb][best_ib] = int(B[best_ib]), int(A[best_ia])
                    verbessert = True
    return gr

beste, beste_kosten = None, float("inf")
alle_kosten = []
for _ in range(VERSUCHE):
    g = optimiere(startloesung())
    k = gesamt_kosten(g)
    alle_kosten.append(k)
    if k < beste_kosten:
        beste, beste_kosten = [grp[:] for grp in g], k

mittlere_kosten = sum(alle_kosten) / len(alle_kosten)
groesste_kosten = max(alle_kosten)

# Referenz: rein zufällige Gruppierung OHNE lokale Optimierung
zufalls_kosten = [gesamt_kosten(startloesung()) for _ in range(1000)]
mittel_zufall = sum(zufalls_kosten) / len(zufalls_kosten)

print(f"\n{'='*50}")
print(f"{N_GRUPPEN} Gruppen, Gesamt-Distanzsumme innerhalb: {beste_kosten:.1f} km (Schnitt/Verein {beste_kosten / M:.1f} km)")
print(f"  (Referenz aus {len(alle_kosten)} optimierten Versuchen: Mittelwert {mittlere_kosten:.1f} km [{mittlere_kosten / M:.1f} km/Verein], "
      f"Maximum {groesste_kosten:.1f} km [{groesste_kosten / M:.1f} km/Verein])")
print(f"  (Reiner Zufall ohne Optimierung, Mittelwert aus {len(zufalls_kosten)} Versuchen: "
      f"{mittel_zufall:.1f} km [{mittel_zufall / M:.1f} km/Verein])\n")
for gi, grp in enumerate(beste, 1):
    n = len(grp)
    summe = gruppen_kosten(grp)
    paare = n * (n - 1) / 2
    avg_paar = summe / paare if paare else 0.0
    avg_verein = summe / n if n else 0.0
    print(f"Gruppe {gi} ({n} Vereine, Distanzsumme {summe:.1f} km, Schnitt/Paar {avg_paar:.1f} km, Schnitt/Verein {avg_verein:.1f} km):")
    print("   " + ", ".join(f"{labels[i]} ({namen[i]})" for i in grp))

ohne = [f"{labels[i]} ({namen[i]})" for i in range(N) if not valid[i]]
if ohne:
    print(f"\nNicht zugeordnet (keine Koordinate): {', '.join(ohne)}")

# Gruppen in Blatt 2
sh2 = out.create_sheet("Gruppen")
sh2.cell(row=1, column=1, value="Gruppe")
sh2.cell(row=1, column=2, value="V.Nr.")
sh2.cell(row=1, column=3, value="Vereinsname")
z = 2
for gi, grp in enumerate(beste, 1):
    for i in grp:
        sh2.cell(row=z, column=1, value=gi)
        sh2.cell(row=z, column=2, value=labels[i])
        sh2.cell(row=z, column=3, value=namen[i])
        z += 1

# Abstandssummen je Gruppe (Übersicht rechts daneben)
sh2.cell(row=1, column=5, value="Gruppe")
sh2.cell(row=1, column=6, value="Anzahl Vereine")
sh2.cell(row=1, column=7, value="Distanzsumme (km)")
sh2.cell(row=1, column=8, value="Ø je Vereinspaar (km)")
sh2.cell(row=1, column=9, value="Ø je Verein (km)")
for gi, grp in enumerate(beste, 1):
    n = len(grp)
    summe = gruppen_kosten(grp)
    paare = n * (n - 1) / 2
    sh2.cell(row=1 + gi, column=5, value=gi)
    sh2.cell(row=1 + gi, column=6, value=n)
    sh2.cell(row=1 + gi, column=7, value=round(summe, 1))
    sh2.cell(row=1 + gi, column=8, value=round(summe / paare, 1) if paare else 0)
    sh2.cell(row=1 + gi, column=9, value=round(summe / n, 1) if n else 0)

for versuch in range(5):
    try:
        out.save(EXCEL_OUT)
        break
    except PermissionError:
        if versuch == 4:
            raise
        time.sleep(0.5)
print(f"\nGespeichert: {EXCEL_OUT} (Matrix + Gruppen)")
if nicht_gefunden:
    print("\nNicht geocodiert:")
    for vnr, adr in nicht_gefunden:
        print(f"  {vnr}: {adr}")
if ungenau:
    print("\nNur ueber Ortsmitte/PLZ aufgeloest (ungenau, keine Strasse oder Pluscode gefunden):")
    for vnr, adr in ungenau:
        print(f"  {vnr}: {adr}")