import io
import math
import os
import random
import re
import time
from collections import Counter

import numpy as np
import openpyxl
import pydeck as pdk
import requests
import streamlit as st
from openlocationcode import openlocationcode as olc
from streamlit_sortables import sort_items

VERSION = "1.02"
OSRM_BATCH = 50   # max Koordinaten pro OSRM-Anfrage; Inter-Chunk-Requests haben 2x davon

st.set_page_config(page_title="GruppenOptimierer", layout="wide")
st.title(f"GruppenOptimierer v{VERSION}")
st.write(
    "Lädt eine Excel-Datei mit Vereinsadressen, berechnet die Fahrtkilometer-Matrix "
    "zwischen den Spielfeldern und teilt die Vereine in ausgewogene Gruppen mit "
    "minimaler interner Fahrtstrecke auf."
)


# ----------------------------------------------------------------------
# Kernlogik (identisch zu mycode.py): Adresse besteht aus "Straße, PLZ Ort"
# oder "Pluscode, PLZ Ort"
# ----------------------------------------------------------------------
PLUSCODE_REGEX = re.compile(r"\b[23456789CFGHJMPQRVWXcfghjmpqrvwx]{2,8}\+[23456789CFGHJMPQRVWXcfghjmpqrvwx]{0,3}\b")


def parse_adresse(adresse):
    # Adresse ist bereinigt auf "Strasse/Platzname, PLZ Ort" bzw. "Pluscode,
    # PLZ Ort" - die Strasse muss daher kein bestimmtes Wort (Weg/Strasse/...)
    # enthalten, sie ist einfach das Segment, das weder PLZ/Ort noch Pluscode ist.
    s = adresse.replace("  ", " ").strip()
    teile = [t.strip() for t in s.split(",") if t.strip()]

    plz_ort = None
    plz_index = None
    for idx, t in enumerate(teile):
        m = re.search(r"\b(\d{5})\s+(.+)", t)
        if m:
            ort = re.split(r"[-]| Zentrum| Innenstadt| Stadtgebiet| Süd| Nord| Ost| West",
                           m.group(2))[0].strip()
            plz_ort = f"{m.group(1)} {ort}"
            plz_index = idx
            break

    pluscode = None
    pluscode_index = None
    for idx, t in enumerate(teile):
        for token in PLUSCODE_REGEX.findall(t):
            if olc.isValid(token):
                pluscode = token.upper()
                pluscode_index = idx
                break
        if pluscode:
            break

    strasse = None
    for idx, t in enumerate(teile):
        if idx == plz_index or idx == pluscode_index:
            continue
        if re.search(r"\boder\b", t, re.I):
            # eher eine unsichere Ortsbeschreibung ("Kemter Wiesen oder Mainzer
            # Str. 199") als eine echte Adresse - nicht als Strasse verwenden.
            continue
        strasse = t
        break

    return strasse, pluscode, plz_ort


def photon_suche(query):
    headers = {"User-Agent": "spielfeld-abstandsmatrix/1.0"}
    r = requests.get("https://photon.komoot.io/api/",
                     params={"q": query, "limit": 1, "lang": "de"},
                     headers=headers, timeout=30)
    r.raise_for_status()
    feats = r.json().get("features", [])
    time.sleep(1.0)
    if feats:
        lon, lat = feats[0]["geometry"]["coordinates"]
        return (float(lat), float(lon))
    return None


def geocode(adresse, cache):
    # quelle gibt an, wie praezise die Koordinate ist: "pluscode"/"strasse"
    # sind genau, "ortsmitte" (nur PLZ/Ort) und "rohtext" sind Fallbacks.
    if adresse in cache:
        return cache[adresse]
    strasse, pluscode, plz_ort = parse_adresse(adresse)
    result = None
    quelle = None

    if pluscode and plz_ort:
        referenz = photon_suche(plz_ort)
        if referenz:
            try:
                voller_code = olc.recoverNearest(pluscode, referenz[0], referenz[1])
                bereich = olc.decode(voller_code)
                result = (bereich.latitudeCenter, bereich.longitudeCenter)
                quelle = "pluscode"
            except Exception:
                result = None

    if result is None:
        varianten = []
        if strasse and plz_ort:
            varianten.append(("strasse", f"{strasse}, {plz_ort}"))
        if plz_ort:
            varianten.append(("ortsmitte", plz_ort))
        if not varianten:
            varianten.append(("rohtext", adresse))
        for label, variante in varianten:
            result = photon_suche(variante)
            if result:
                quelle = label
                break

    cache[adresse] = (result, quelle)
    return cache[adresse]


def osrm_matrix_chunked(gueltige_coords, on_block=None):
    """Berechnet MxM-Distanzmatrix in Batches (umgeht URL-Laengenlimit bei >OSRM_BATCH Coords)."""
    M = len(gueltige_coords)
    if M <= OSRM_BATCH:
        coord_str = ";".join(f"{lon:.6f},{lat:.6f}" for lat, lon in gueltige_coords)
        r = requests.get(f"https://router.project-osrm.org/table/v1/driving/{coord_str}",
                         params={"annotations": "distance"}, timeout=120)
        r.raise_for_status()
        osrm = r.json()["distances"]
        if on_block:
            on_block(1, 1)
        return [[round(osrm[a][b] / 1000, 1) if osrm[a][b] is not None else None
                 for b in range(M)] for a in range(M)]

    raw = [[None] * M for _ in range(M)]
    chunks = [list(range(s, min(s + OSRM_BATCH, M))) for s in range(0, M, OSRM_BATCH)]
    n_chunks = len(chunks)
    n_tasks = n_chunks * (n_chunks + 1) // 2
    fertig = 0
    for ci, ca in enumerate(chunks):
        for cj in range(ci, n_chunks):
            cb = chunks[cj]
            combined = ca if ci == cj else ca + cb
            n_a = len(ca)
            coord_str = ";".join(f"{gueltige_coords[k][1]:.6f},{gueltige_coords[k][0]:.6f}"
                                 for k in combined)
            r = requests.get(f"https://router.project-osrm.org/table/v1/driving/{coord_str}",
                             params={"annotations": "distance"}, timeout=120)
            r.raise_for_status()
            osrm = r.json()["distances"]
            if ci == cj:
                for a, i in enumerate(ca):
                    for b, j in enumerate(ca):
                        d = osrm[a][b]
                        raw[i][j] = round(d / 1000, 1) if d is not None else None
            else:
                for a, i in enumerate(ca):
                    for b, j in enumerate(cb):
                        d_ab = osrm[a][n_a + b]
                        d_ba = osrm[n_a + b][a]
                        raw[i][j] = round(d_ab / 1000, 1) if d_ab is not None else None
                        raw[j][i] = round(d_ba / 1000, 1) if d_ba is not None else None
            fertig += 1
            if on_block:
                on_block(fertig, n_tasks)
    return raw


def baue_geocoding_log(labels, namen, adressen, quellen, valid):
    ortsmitte = [(labels[i], namen[i], adressen[i]) for i in range(len(labels))
                 if quellen[i] in ("ortsmitte", "rohtext")]
    fehlend = [(labels[i], namen[i], adressen[i]) for i in range(len(labels)) if not valid[i]]

    zeilen = []
    if ortsmitte:
        zeilen.append("Nur ueber Ortsmitte/PLZ aufgeloest (ungenau, keine Strasse oder Pluscode gefunden):")
        for vnr, name, adr in ortsmitte:
            zeilen.append(f"  {vnr} ({name}): {adr}")
        zeilen.append("")
    if fehlend:
        zeilen.append("Keine Koordinate gefunden:")
        for vnr, name, adr in fehlend:
            zeilen.append(f"  {vnr} ({name}): {adr or '(keine Adresse)'}")
        zeilen.append("")
    if not zeilen:
        zeilen.append("Alle Adressen wurden ueber Strasse oder Pluscode praezise aufgeloest.")
    zeilen.append("")
    zeilen.append("Geocoding-Dienst: Photon von Komoot – https://photon.komoot.io")
    return "\n".join(zeilen)


def dist(matrix, i, j):
    a, b = matrix[i][j], matrix[j][i]
    if a is None and b is None:
        return 0.0
    if a is None:
        return b
    if b is None:
        return a
    return (a + b) / 2


def gruppen_kosten(matrix, m):
    s = 0.0
    for x in range(len(m)):
        for y in range(x + 1, len(m)):
            s += dist(matrix, m[x], m[y])
    return s


def gesamt_kosten(matrix, gr):
    return sum(gruppen_kosten(matrix, g) for g in gr)


GRUPPEN_EMOJI = ["🔵", "🟢", "🟣", "🟠", "🟡", "🟤", "⚫", "⚪"]


def gruppen_emoji(gi):
    return GRUPPEN_EMOJI[gi % len(GRUPPEN_EMOJI)]


def item_label(labels, namen, i, urspruenglich_gruppe):
    # Symbol zeigt die Gruppe im Optimalzustand - steht direkt im Text, damit
    # es beim Verschieben mit dem Verein "mitwandert" (positionsbasiertes
    # CSS-Styling ist bei dieser Komponente nicht zuverlaessig, da sich die
    # DOM-Reihenfolge beim Draggen aendert).
    emoji = gruppen_emoji(urspruenglich_gruppe[i])
    return f"{emoji} {labels[i]} ({namen[i]})"


def kanonisch(gruppen):
    # Reihenfolge innerhalb einer Gruppe ist irrelevant fuer die Kosten -
    # feste Sortierung verhindert, dass reines Umsortieren als Aenderung gilt.
    return [sorted(grp) for grp in gruppen]


GRUPPEN_RGB = [
    [37, 99, 235], [22, 163, 74], [147, 51, 234], [234, 88, 12],
    [8, 145, 178], [202, 138, 4], [219, 39, 119], [77, 124, 15],
]


def gruppen_rgb(gi):
    return GRUPPEN_RGB[gi % len(GRUPPEN_RGB)]


def baue_ausgabe_excel(labels, namen, matrix, gruppen):
    out = openpyxl.Workbook()
    sh = out.active
    sh.title = "Abstandsmatrix_km"
    sh.cell(row=1, column=1, value="V.Nr.")
    for j, lab in enumerate(labels):
        sh.cell(row=1, column=2 + j, value=lab)
    for i, lab in enumerate(labels):
        sh.cell(row=2 + i, column=1, value=lab)
        for j in range(len(labels)):
            sh.cell(row=2 + i, column=2 + j, value=matrix[i][j])

    sh2 = out.create_sheet("Gruppen")
    sh2.cell(row=1, column=1, value="Gruppe")
    sh2.cell(row=1, column=2, value="V.Nr.")
    sh2.cell(row=1, column=3, value="Vereinsname")
    z = 2
    for gi, grp in enumerate(gruppen, 1):
        for i in grp:
            sh2.cell(row=z, column=1, value=gi)
            sh2.cell(row=z, column=2, value=labels[i])
            sh2.cell(row=z, column=3, value=namen[i])
            z += 1

    sh2.cell(row=1, column=5, value="Gruppe")
    sh2.cell(row=1, column=6, value="Anzahl Vereine")
    sh2.cell(row=1, column=7, value="Distanzsumme (km)")
    sh2.cell(row=1, column=8, value="Ø je Vereinspaar (km)")
    sh2.cell(row=1, column=9, value="Ø je Verein (km)")
    for gi, grp in enumerate(gruppen, 1):
        n = len(grp)
        summe = sum(dist(matrix, a, b) for ai, a in enumerate(grp) for b in grp[ai + 1:])
        paare = n * (n - 1) / 2
        sh2.cell(row=1 + gi, column=5, value=gi)
        sh2.cell(row=1 + gi, column=6, value=n)
        sh2.cell(row=1 + gi, column=7, value=round(summe, 1))
        sh2.cell(row=1 + gi, column=8, value=round(summe / paare, 1) if paare else 0)
        sh2.cell(row=1 + gi, column=9, value=round(summe / n, 1) if n else 0)

    buf = io.BytesIO()
    out.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------
with st.sidebar:
    st.header("Eingabe")
    datei = st.file_uploader("Meldelisten-Excel (.xlsx)", type=["xlsx"])
    sheet = st.text_input("Worksheet-Name (leer = erstes Worksheet)", value="")
    n_gruppen_input = st.number_input("Anzahl Gruppen", min_value=2, max_value=100, value=4, step=1)
    with st.expander("Erweiterte Einstellungen"):
        seed = st.number_input("Zufallssaat", value=42, step=1)
        versuche = st.number_input("Zufallsneustarts (Heuristik)", min_value=1, max_value=200, value=40, step=1)
    start = st.button("Berechnen", type="primary", disabled=datei is None)

if not datei:
    st.info("Bitte links eine Excel-Datei hochladen und die Anzahl der Gruppen wählen.")
    st.stop()

if start:
    # ------------------------------------------------------------------
    # 1) Excel einlesen: Spalte A = V.Nr., Spalte B = Vereinsname, Spalte L = Adresse
    # ------------------------------------------------------------------
    wb = openpyxl.load_workbook(datei, data_only=True)
    sheet_name = sheet.strip() or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        st.error(f"Worksheet '{sheet_name}' nicht gefunden. Vorhanden: {', '.join(wb.sheetnames)}")
        st.stop()
    ws = wb[sheet_name]

    vereine = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vnr = row[0]
        name = row[1] if len(row) > 1 else None
        adr = row[11] if len(row) > 11 else None
        if vnr is None:
            continue
        name = str(name).strip() if name else ""
        adr = str(adr).strip() if adr else ""
        vereine.append((str(vnr), name, adr))

    labels = [vnr for vnr, _, _ in vereine]
    namen = [name for _, name, _ in vereine]
    N = len(vereine)
    n_gruppen = int(n_gruppen_input)
    st.write(f"**{N} Vereine eingelesen (Worksheet: {sheet_name}).**")

    if N < n_gruppen:
        st.error(f"Nur {N} Vereine, aber {n_gruppen} Gruppen angefordert.")
        st.stop()

    # ------------------------------------------------------------------
    # 2)+3) Geocoding (Photon)
    # ------------------------------------------------------------------
    cache = {}
    coords = []
    quellen = []
    nicht_gefunden = []
    progress = st.progress(0.0, text="Geocoding...")
    for idx, (vnr, _, adr) in enumerate(vereine):
        if not adr:
            coords.append(None)
            quellen.append(None)
            nicht_gefunden.append((vnr, "(keine Adresse)"))
        else:
            c, q = geocode(adr, cache)
            coords.append(c)
            quellen.append(q)
            if c is None:
                nicht_gefunden.append((vnr, adr))
        progress.progress((idx + 1) / N, text=f"Geocoding {idx + 1}/{N}: {vnr}")
    progress.empty()

    valid = [c is not None for c in coords]
    gueltig = [i for i in range(N) if valid[i]]
    M = len(gueltig)

    if M < n_gruppen:
        st.error(f"Nur {M} Vereine konnten geocodiert werden, aber {n_gruppen} Gruppen angefordert.")
        st.stop()

    # Nebenbedingung: Vereine mit gleicher Vereinsnummer in verschiedene Gruppen
    _id_map = {}
    for i in gueltig:
        _id_map.setdefault(labels[i], []).append(i)
    mehrfach = {vnr: idxs for vnr, idxs in _id_map.items() if len(idxs) > 1}
    if mehrfach:
        _max = max(len(v) for v in mehrfach.values())
        if _max > n_gruppen:
            st.error(
                f"Eine Vereinsnummer erscheint {_max}× – die Nebenbedingung (gleiche "
                f"Vereinsnummer in verschiedene Gruppen) kann mit nur {n_gruppen} Gruppen "
                f"nicht erfüllt werden. Bitte Gruppenanzahl erhöhen."
            )
            st.stop()

    # ------------------------------------------------------------------
    # 4) Fahrtkilometer-Matrix (OSRM, in Batches fuer grosse Vereinsmengen)
    # ------------------------------------------------------------------
    gueltige_coords = [coords[i] for i in gueltig]
    progress_osrm = st.progress(0.0, text=f"Berechne Fahrtkilometer-Matrix (OSRM, {M} Vereine)...")
    raw = osrm_matrix_chunked(
        gueltige_coords,
        on_block=lambda f, g: progress_osrm.progress(f / g,
                                                      text=f"OSRM-Matrix: Block {f}/{g}"),
    )
    progress_osrm.empty()

    matrix = [[None] * N for _ in range(N)]
    for a, i in enumerate(gueltig):
        for b, j in enumerate(gueltig):
            matrix[i][j] = raw[a][b]

    # ------------------------------------------------------------------
    # 5) In n_gruppen ausgewogene Gruppen aufteilen (minimale interne Distanz)
    # ------------------------------------------------------------------
    random.seed(int(seed))
    basis, rest = M // n_gruppen, M % n_gruppen
    groessen = [basis + (1 if g < rest else 0) for g in range(n_gruppen)]

    # Symmetrische Distanzmatrix fuer vektorisierte Optimierung
    sym = np.zeros((N, N))
    for _i in range(N):
        for _j in range(N):
            _a, _b = matrix[_i][_j], matrix[_j][_i]
            if _a is not None and _b is not None:
                sym[_i, _j] = (_a + _b) / 2.0
            elif _a is not None:
                sym[_i, _j] = float(_a)
            elif _b is not None:
                sym[_i, _j] = float(_b)

    def startloesung():
        pool = gueltig[:]
        random.shuffle(pool)
        if not mehrfach:
            gr, pos = [], 0
            for g in range(n_gruppen):
                gr.append(pool[pos:pos + groessen[g]])
                pos += groessen[g]
            return gr
        # Nebenbedingung: gleiche Vereinsnummer nicht in derselben Gruppe
        gr = [[] for _ in range(n_gruppen)]
        kapaz = groessen[:]
        gr_ids = [set() for _ in range(n_gruppen)]
        for idx in pool:
            lid = labels[idx]
            frei = [g for g in range(n_gruppen) if kapaz[g] > 0 and lid not in gr_ids[g]]
            if not frei:
                frei = [g for g in range(n_gruppen) if kapaz[g] > 0]
            g = random.choice(frei)
            gr[g].append(idx)
            kapaz[g] -= 1
            gr_ids[g].add(lid)
        return gr

    def optimiere(gr):
        # Vektorisierte Delta-Formel: delta[ia,ib] = sA[b]-sA[a] + sB[a]-sB[b] - 2*sym[a,b]
        # Infeasible Tausche (wuerden Nebenbedingung verletzen) erhalten delta=inf.
        verbessert = True
        while verbessert:
            verbessert = False
            for ga in range(n_gruppen):
                for gb in range(ga + 1, n_gruppen):
                    A = np.array(gr[ga])
                    B = np.array(gr[gb])
                    sA = sym[:, A].sum(axis=1)
                    sB = sym[:, B].sum(axis=1)
                    delta = (sA[B][np.newaxis, :] - sA[A][:, np.newaxis]
                           + sB[A][:, np.newaxis] - sB[B][np.newaxis, :]
                           - 2.0 * sym[np.ix_(A, B)])
                    if mehrfach:
                        lA = np.array([labels[int(k)] for k in A])
                        lB = np.array([labels[int(k)] for k in B])
                        cnt_ga = Counter(labels[int(k)] for k in gr[ga])
                        cnt_gb = Counter(labels[int(k)] for k in gr[gb])
                        cnt_B_in_ga = np.array([cnt_ga[lb] for lb in lB])
                        cnt_A_in_gb = np.array([cnt_gb[la] for la in lA])
                        same = (lA[:, np.newaxis] == lB[np.newaxis, :])
                        # Nach Tausch: B[ib] geht in ga → verletzt NB wenn B-Label noch in ga
                        eff_ga = cnt_B_in_ga[np.newaxis, :] - same.astype(int)
                        # Nach Tausch: A[ia] geht in gb → verletzt NB wenn A-Label noch in gb
                        eff_gb = cnt_A_in_gb[:, np.newaxis] - same.astype(int)
                        delta[(eff_ga > 0) | (eff_gb > 0)] = np.inf
                    best = int(delta.argmin())
                    best_ia, best_ib = divmod(best, len(B))
                    if delta.flat[best] < -1e-9:
                        gr[ga][best_ia], gr[gb][best_ib] = int(B[best_ib]), int(A[best_ia])
                        verbessert = True
        return gr

    with st.spinner(f"Optimiere Gruppen ({int(versuche)} Zufallsneustarts)..."):
        beste, beste_kosten = None, float("inf")
        alle_kosten = []
        for _ in range(int(versuche)):
            g = optimiere(startloesung())
            k = gesamt_kosten(matrix, g)
            alle_kosten.append(k)
            if k < beste_kosten:
                beste, beste_kosten = [grp[:] for grp in g], k

    mittlere_kosten = sum(alle_kosten) / len(alle_kosten)
    groesste_kosten = max(alle_kosten)

    with st.spinner("Berechne Referenz (zufällige Gruppierung ohne Optimierung)..."):
        zufalls_kosten = [gesamt_kosten(matrix, startloesung()) for _ in range(1000)]
    mittel_zufall = sum(zufalls_kosten) / len(zufalls_kosten)
    verbesserung = (mittel_zufall - beste_kosten) / mittel_zufall * 100

    st.session_state.ergebnis = dict(
        labels=labels, namen=namen, matrix=matrix, valid=valid, N=N, M=M,
        n_gruppen=n_gruppen, beste=beste, beste_kosten=beste_kosten,
        alle_kosten_n=len(alle_kosten), mittlere_kosten=mittlere_kosten,
        groesste_kosten=groesste_kosten, zufalls_kosten_n=len(zufalls_kosten),
        mittel_zufall=mittel_zufall, verbesserung=verbesserung,
        dateiname=datei.name, coords=coords, quellen=quellen,
        adressen=[adr for _, _, adr in vereine],
        mehrfach_ids=[(vnr, len(idxs)) for vnr, idxs in mehrfach.items()],
    )
    st.session_state.aktuelle_gruppen = kanonisch(beste)
    st.session_state.vorherige_gruppen = kanonisch(beste)
    st.session_state.sortable_version = st.session_state.get("sortable_version", 0) + 1

if "ergebnis" not in st.session_state:
    st.info("Bitte 'Berechnen' klicken, um Matrix und Gruppen zu berechnen.")
    st.stop()

erg = st.session_state.ergebnis
labels = erg["labels"]
namen = erg["namen"]
matrix = erg["matrix"]
valid = erg["valid"]
N = erg["N"]
M = erg["M"]
n_gruppen = erg["n_gruppen"]
beste = erg["beste"]
beste_kosten = erg["beste_kosten"]
coords = erg["coords"]
quellen = erg["quellen"]
adressen = erg["adressen"]
datei_basis, datei_ext = os.path.splitext(erg["dateiname"])

# ----------------------------------------------------------------------
# Referenzwerte (Optimum, Zufallsneustarts, reiner Zufall)
# ----------------------------------------------------------------------
st.success(
    f"{n_gruppen} Gruppen, optimale Distanzsumme: {beste_kosten:.1f} km "
    f"(Schnitt/Verein {beste_kosten / M:.1f} km)"
)
st.caption(
    f"Referenz aus {erg['alle_kosten_n']} optimierten Versuchen: "
    f"Mittelwert {erg['mittlere_kosten']:.1f} km, Maximum {erg['groesste_kosten']:.1f} km"
)
st.caption(
    f"Reiner Zufall ohne Optimierung, Mittelwert aus {erg['zufalls_kosten_n']} Versuchen: "
    f"{erg['mittel_zufall']:.1f} km [{erg['mittel_zufall'] / M:.1f} km/Verein] "
    f"-> Verbesserung durch Optimierung: {erg['verbesserung']:.0f} %"
)

mehrfach_ids = erg.get("mehrfach_ids", [])
if mehrfach_ids:
    details = ", ".join(f"{vnr} ({n}×)" for vnr, n in mehrfach_ids)
    st.info(
        f"Nebenbedingung aktiv: {len(mehrfach_ids)} Vereinsnummer(n) erscheinen mehrfach "
        f"und wurden auf verschiedene Gruppen verteilt. ({details})"
    )

ohne = [f"{labels[i]} ({namen[i]})" for i in range(N) if not valid[i]]
if ohne:
    st.warning("Nicht zugeordnet (keine Koordinate): " + ", ".join(ohne))

ungenau_n = sum(1 for q in quellen if q in ("ortsmitte", "rohtext"))
if ungenau_n:
    st.warning(
        f"{ungenau_n} Adresse(n) konnten nicht praezise aufgeloest werden und wurden nur "
        "ueber die Ortsmitte (PLZ/Ort) geocodiert. Details siehe Geocoding-Log."
    )
st.download_button(
    "Geocoding-Log herunterladen",
    data=baue_geocoding_log(labels, namen, adressen, quellen, valid).encode("utf-8"),
    file_name=f"{datei_basis}-Geocoding-Log.txt",
    mime="text/plain",
)

# ----------------------------------------------------------------------
# Gruppen interaktiv anpassen (Drag & Drop)
# ----------------------------------------------------------------------
st.subheader("Gruppen anpassen")
st.caption("Vereine per Drag & Drop zwischen Gruppen verschieben. "
           "Das Symbol zeigt die Gruppe im Optimalzustand, unabhängig von der aktuellen Position.")
st.caption(" · ".join(
    f"{gruppen_emoji(gi)} Gruppe {gi + 1}" for gi in range(n_gruppen)
))

st.caption("\"Ggü. vorherigem Zustand\" vergleicht immer mit dem zuletzt gespeicherten Vergleichspunkt "
           "(nicht automatisch mit dem letzten einzelnen Zug) - so lassen sich mehrere Änderungen am Stück bewerten.")
knopf1, knopf2 = st.columns(2)
if knopf1.button("Zurücksetzen auf Optimum"):
    st.session_state.aktuelle_gruppen = kanonisch(beste)
    st.session_state.vorherige_gruppen = kanonisch(beste)
    st.session_state.sortable_version += 1
    st.rerun()
if knopf2.button("Aktuellen Zustand als Vergleichspunkt speichern"):
    st.session_state.vorherige_gruppen = kanonisch(st.session_state.aktuelle_gruppen)

aktuelle = st.session_state.aktuelle_gruppen

urspruenglich_gruppe = {}
for gi, grp in enumerate(beste):
    for idx in grp:
        urspruenglich_gruppe[idx] = gi

label_zu_idx = {item_label(labels, namen, i, urspruenglich_gruppe): i for i in urspruenglich_gruppe}

sortable_input = [
    {"header": f"Gruppe {gi + 1}", "items": [item_label(labels, namen, i, urspruenglich_gruppe) for i in grp]}
    for gi, grp in enumerate(aktuelle)
]

sortable_output = sort_items(
    sortable_input,
    multi_containers=True,
    direction="vertical",
    custom_style=".sortable-item, .sortable-item:hover { background-color: #475569 !important; color: #fff !important; text-align: left !important; }",
    key=f"gruppen_sortable_{st.session_state.sortable_version}",
)

neue_gruppen = kanonisch([[label_zu_idx[lbl] for lbl in container["items"]] for container in sortable_output])

if neue_gruppen != aktuelle:
    st.session_state.aktuelle_gruppen = neue_gruppen

aktuelle = st.session_state.aktuelle_gruppen
vorherige = st.session_state.vorherige_gruppen

# Nebenbedingung pruefen (nur Warnung, kein Blockieren)
if mehrfach_ids:
    nb_verletzungen = []
    for gi, grp in enumerate(aktuelle):
        seen = set()
        for idx in grp:
            vnr = labels[idx]
            if vnr in seen:
                nb_verletzungen.append((gi + 1, vnr))
            seen.add(vnr)
    if nb_verletzungen:
        details = "; ".join(f"Gruppe {gi}: V.Nr. {vnr}" for gi, vnr in nb_verletzungen)
        st.warning(f"Nebenbedingung verletzt: gleiche Vereinsnummer in einer Gruppe – {details}")

aktuelle_kosten = gesamt_kosten(matrix, aktuelle)
vorherige_kosten = gesamt_kosten(matrix, vorherige)
delta_optimal = aktuelle_kosten - beste_kosten
delta_vorher = aktuelle_kosten - vorherige_kosten

col1, col2, col3 = st.columns(3)
col1.metric("Distanzsumme aktuell", f"{aktuelle_kosten:.1f} km")
col2.metric("Ggü. Optimum", f"{delta_optimal:+.1f} km",
            delta=f"{delta_optimal:+.1f} km", delta_color="inverse")
col3.metric("Ggü. vorherigem Zustand", f"{delta_vorher:+.1f} km",
            delta=f"{delta_vorher:+.1f} km", delta_color="inverse")

gruppen_spalten = st.columns(n_gruppen)
for gi in range(n_gruppen):
    grp_akt = aktuelle[gi]
    grp_opt = beste[gi]
    grp_vor = vorherige[gi]
    n = len(grp_akt)
    summe_akt = gruppen_kosten(matrix, grp_akt)
    summe_opt = gruppen_kosten(matrix, grp_opt)
    summe_vor = gruppen_kosten(matrix, grp_vor)
    paare = n * (n - 1) / 2
    avg_paar = summe_akt / paare if paare else 0.0
    avg_verein = summe_akt / n if n else 0.0
    delta_opt_gi = summe_akt - summe_opt
    delta_vor_gi = summe_akt - summe_vor
    with gruppen_spalten[gi]:
        st.markdown(f"**Gruppe {gi + 1}** ({n} Vereine)")
        st.metric("km intern", f"{summe_akt:.1f} km")
        st.caption(f"Ø/Paar: {avg_paar:.1f} km")
        st.caption(f"Ø/Verein: {avg_verein:.1f} km")
        st.metric("Ggü. Optimum", f"{delta_opt_gi:+.1f} km",
                  delta=f"{delta_opt_gi:+.1f} km", delta_color="inverse")
        st.metric("Ggü. vorherigem Zustand", f"{delta_vor_gi:+.1f} km",
                  delta=f"{delta_vor_gi:+.1f} km", delta_color="inverse")

# ----------------------------------------------------------------------
# Karte
# ----------------------------------------------------------------------
if "zeige_karte" not in st.session_state:
    st.session_state.zeige_karte = True
karte_label = "Karte ausblenden" if st.session_state.zeige_karte else "Gruppen auf Karte anzeigen"
if st.button(karte_label, key="karte_toggle"):
    st.session_state.zeige_karte = not st.session_state.zeige_karte
    st.rerun()

if st.session_state.zeige_karte:
    aktuelle_gruppe_von = {}
    for gi, grp in enumerate(aktuelle):
        for idx in grp:
            aktuelle_gruppe_von[idx] = gi

    punkte = []
    for idx, gi in aktuelle_gruppe_von.items():
        if coords[idx] is None:
            continue
        lat, lon = coords[idx]
        punkte.append({
            "lat": lat, "lon": lon,
            "farbe": gruppen_rgb(gi),
            "label": f"{labels[idx]} ({namen[idx]})",
            "gruppe": gi + 1,
        })

    # Mehrere Vereine koennen exakt denselben geocodierten Punkt haben
    # (z.B. gleiche Ortsmitte als Fallback) - leicht auffaechern, damit
    # sich die Marker nicht gegenseitig verdecken.
    orte = {}
    for p in punkte:
        orte.setdefault((p["lat"], p["lon"]), []).append(p)
    for gruppe_am_ort in orte.values():
        n = len(gruppe_am_ort)
        if n > 1:
            radius_grad = 0.0015
            for k, p in enumerate(gruppe_am_ort):
                winkel = 2 * math.pi * k / n
                p["lat"] += radius_grad * math.sin(winkel)
                p["lon"] += radius_grad * math.cos(winkel) / math.cos(math.radians(p["lat"]))

    if punkte:
        st.caption(" · ".join(
            f'<span style="color:rgb({gruppen_rgb(gi)[0]},{gruppen_rgb(gi)[1]},{gruppen_rgb(gi)[2]})">⬤</span> '
            f"Gruppe {gi + 1} (aktuell)"
            for gi in range(n_gruppen)
        ), unsafe_allow_html=True)

        schicht = pdk.Layer(
            "ScatterplotLayer",
            data=punkte,
            get_position="[lon, lat]",
            get_fill_color="farbe",
            get_radius=400,
            radius_min_pixels=6,
            radius_max_pixels=20,
            stroked=True,
            get_line_color=[255, 255, 255],
            line_width_min_pixels=1,
            pickable=True,
        )
        ansicht = pdk.data_utils.compute_view([[p["lon"], p["lat"]] for p in punkte])
        ansicht.zoom = max(0, ansicht.zoom - 0.6)  # Sicherheitsabstand, damit Randpunkte nicht abgeschnitten werden
        st.pydeck_chart(pdk.Deck(
            layers=[schicht],
            initial_view_state=ansicht,
            tooltip={"html": "<b>{label}</b><br/>Gruppe {gruppe}"},
        ), height=600)
    else:
        st.info("Keine Koordinaten zum Anzeigen vorhanden.")

# ----------------------------------------------------------------------
# Download
# ----------------------------------------------------------------------
excel_bytes = baue_ausgabe_excel(labels, namen, matrix, aktuelle)
ausgabe_dateiname = f"{datei_basis}-Gruppeneinteilung{datei_ext or '.xlsx'}"
st.download_button(
    f"{ausgabe_dateiname} herunterladen",
    data=excel_bytes,
    file_name=ausgabe_dateiname,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
